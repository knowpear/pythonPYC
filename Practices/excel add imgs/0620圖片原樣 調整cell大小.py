import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# 定义路径
excel_path = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\(营业厅物品汇总)表格视图.xlsx"
images_folder = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\assembly"
output_path = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\output.xlsx"

# 加载Excel工作簿和工作表
wb = load_workbook(excel_path)
ws = wb.active

# 遍历所有单元格，查找包含图片名称的单元格
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            image_name = cell.value
            image_path = os.path.join(images_folder, image_name)
            if os.path.exists(image_path):
                # 插入图像到Excel单元格
                img = Image(image_path)
                ws.add_image(img, cell.coordinate)

                # 获取图片尺寸并调整单元格大小
                img_width, img_height = img.width, img.height
                col_letter = cell.column_letter
                row_index = cell.row

                # 假设默认的宽度和高度为0.15英寸（单位：英寸）
                default_col_width = 0.15 * 10  # 默认列宽度乘以 10
                default_row_height = 0.15 * 72  # 默认行高度乘以 72

                # 计算列宽和行高
                new_col_width = img_width / 7.5  # 7.5是经验值，可以根据需要调整
                new_row_height = img_height * 0.75  # 0.75是经验值，可以根据需要调整

                # 更新列宽和行高
                ws.column_dimensions[col_letter].width = max(new_col_width, ws.column_dimensions[col_letter].width)
                ws.row_dimensions[row_index].height = max(new_row_height, ws.row_dimensions[row_index].height)

# 保存新的Excel文件
wb.save(output_path)

print(f"完成! 输出文件保存在: {output_path}")