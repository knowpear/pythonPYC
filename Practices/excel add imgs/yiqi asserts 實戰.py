import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import os


def insert_images_to_excel(excel_path, images_folder):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # 遍历所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            # 检查单元格是否有图片名称
            if cell.value and isinstance(cell.value, str):
                image_names = cell.value.splitlines()  # 以换行符分隔获取图片名称列表

                # 计算总宽度并调整单元格宽度
                total_width = len(image_names) * 15  # 调整总宽度
                sheet.column_dimensions[cell.column_letter].width = total_width  # 设置单元格宽度

                for index, image_name in enumerate(image_names):
                    image_name = image_name.strip()  # 去除空格
                    image_path = os.path.join(images_folder, image_name)

                    # 检查图片文件是否存在
                    if os.path.exists(image_path):
                        # 加载图片并调整大小
                        with PILImage.open(image_path) as img:
                            img.thumbnail((100, 100), PILImage.LANCZOS)  # 调整图片大小，保持比例
                            resized_image_path = os.path.join(images_folder, f"resized_{image_name}")
                            img.save(resized_image_path)  # 保存调整后的图片

                            # 插入图片到单元格
                            excel_img = Image(resized_image_path)

                            # 计算每张图片的水平偏移量
                            offset_x = index * 100  # 100个像素的水平间距，避免重叠

                            # 设置图片位置，考虑水平偏移量
                            col = cell.column_letter
                            row = cell.row
                            # 计算新列的字母
                            new_col_index = openpyxl.utils.cell.column_index_from_string(col) + index
                            new_col = openpyxl.utils.cell.get_column_letter(new_col_index)
                            excel_img.anchor = f'{new_col}{row}'
                            sheet.add_image(excel_img)
                    else:
                        print(f"图片 '{image_name}' 不存在于目录 '{images_folder}' 中")

    # 保存修改后的 Excel 文件
    wb.save(output_path)

# 设置 Excel 文件路径和图片文件夹路径
excel_path = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\(营业厅物品汇总)表格视图.xlsx"
images_folder = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\assembly"
output_path = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\output.xlsx"

insert_images_to_excel(excel_path, images_folder)
