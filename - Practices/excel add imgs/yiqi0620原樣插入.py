import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# 定义路径
excel_path = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\(营业厅物品汇总)表格视图.xlsx"
images_folder = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\assembly"
output_path = r"C:\Users\DaiYi\Desktop\营业厅资产统计(202406051240)\output.xlsx"

# 加载Excel工作簿和工作表
wb = load_workbook(excel_path)
ws = wb.active

# 遍历所有单元格，查找包含图片名称的单元格
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            image_name = cell.value
            image_path = os.path.join(images_folder, image_name)
            if os.path.exists(image_path):
                # 插入图像到Excel单元格
                img = Image(image_path)
                ws.add_image(img, cell.coordinate)

# 保存新的Excel文件
wb.save(output_path)

print(f"完成! 输出文件保存在: {output_path}")
